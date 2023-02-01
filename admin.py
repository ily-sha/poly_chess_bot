from main import *
import main
import openpyxl
import json
import requests
import datetime
import io
from params import Params
from String import *
admin_tools = [[str_get_attendance_adm], [mark_student_command], [add_new_student, add_lesson_date],
               [add_new_semester_command, change_list]]
admins = ["peterthegreat", "echecs95"]
process_admin = {}


def get_all_dates():
    book = openpyxl.load_workbook(filename)
    sheet = book[main.sheetname]
    all_dates = []
    for row in sheet.iter_rows():
        for title in row:
            if isinstance(title.value, datetime.datetime):
                date = f"{title.value.day}-{title.value.month}-{title.value.year}"
                all_dates.append(date)
        return all_dates


def add_new_person_params(chat_id, text):
    person = process_admin[chat_id]
    if person.get_surname() is None:
        person.set_surname(text)
        send_message(chat_id, Params(text=choose_course,
                                     reply_markup=create_reply_markup([[first_course, second_course]])))
    else:
        if person.get_course() is None:
            person.set_course(int(text.split()[0]))
            send_message(chat_id, Params(text=choose_advantages,
                                         reply_markup=create_reply_markup([[no_advantage]])))
        else:
            if text != no_advantage:
                person.set_advantage(text)
            book = openpyxl.load_workbook(filename)
            sheet = book[main.sheetname]
            new_row = sheet.max_row + 1
            sheet[course_column + str(new_row)] = int(person.get_course())
            sheet[surname_column + str(new_row)] = person.get_surname()
            if person.get_advantage() is not None:
                sheet[advantage_column + str(new_row)] = person.get_advantage()
            book.save(filename)
            send_message(chat_id,
                         Params(text=f"Ученик/ца {person.get_surname()} {person.get_course()}-ого курса добавлен/a",
                                reply_markup=create_reply_markup(admin_tools)))
            process_admin.pop(chat_id)


def send_media(chat_id, filename, type):
    data = {
        "chat_id": chat_id,
        "media": json.dumps([
            {"type": type, "media": "attach://file"}
        ])
    }
    file = {
        "file": open(filename, 'rb')
    }
    requests.post(f"{URL + TOKEN}/sendMediaGroup", data=data, files=file)


def get_new_book(message):
    result = requests.get(f"{URL + TOKEN}/getFile?file_id={message.get_file_id()}").json()
    xlsx = io.BytesIO(requests.get(f"https://api.telegram.org/file/bot{TOKEN}/{result['result']['file_path']}").content)
    new_book = openpyxl.load_workbook(xlsx)
    main.sheetname = new_book.sheetnames[0]
    new_sheet = new_book[main.sheetname]
    new_data = new_sheet[new_sheet.dimensions]
    return new_data


def put_dates(main_sheet):
    next_day = datetime.date.today()
    while next_day.weekday() != 5:
        next_day += datetime.timedelta(days=1)
    next_column = find_next_column(main_sheet.max_column)
    for i in range(16):
        main_sheet[next_column + "1"].value = next_day
        next_day += datetime.timedelta(days=7)
        next_column = find_next_column(main_sheet.max_column)


def admin_method(message):
    chat_id = message.get_chat_id()
    if message.is_text_message():
        text = message.get_text()
        if text == start_command:
            process_admin.clear()
            send_message(chat_id, Params(text=administrator_permission,
                                         reply_markup=create_reply_markup(admin_tools)))
        elif text in openpyxl.load_workbook(filename).sheetnames:
            main.sheetname = text
            send_message(chat_id, Params(text=f"Актуальный лист - {text}",
                                         reply_markup=create_reply_markup(admin_tools)))
        elif text == str_get_attendance_adm:
            send_media(chat_id, filename, "document")
        elif text == add_lesson_date:
            send_message(chat_id, Params(text=write_correct_date))
            process_admin[chat_id] = add_lesson_date
        elif text == mark_student_command:
            send_message(chat_id, Params(text=choose_course,
                                         reply_markup=create_reply_markup([[first_course, second_course]])))
        elif text == add_new_semester_command:
            send_message(chat_id, Params(text=requirements_of_sheet))
            send_media(chat_id, requirements_of_sheet_file, "photo")
        elif text == change_list:
            book = openpyxl.load_workbook(filename)

            send_message(chat_id, Params(text=choose_sheet,
                                         reply_markup=create_reply_markup(chunked_list(book.sheetnames))))
        elif is_surname_exist(message.get_text()):
            process_admin[chat_id] = [text, choose_date]
            today = f"{str(datetime.date.today().day)}-{str(datetime.date.today().month)}-{str(datetime.date.today().year)}"
            arr = chunked_list(get_all_dates())
            arr.insert(0, [today])
            send_message(chat_id, Params(text=choose_date,
                                         reply_markup=create_reply_markup(arr)))
        elif text == add_new_student:
            send_message(chat_id, Params(text=write_student_surname))
            process_admin[chat_id] = NewPerson()
        elif chat_id in process_admin.keys() and isinstance(process_admin[chat_id], NewPerson):
            add_new_person_params(chat_id, text)
        elif course in text:
            choose_surname(chat_id, text)
        elif chat_id in process_admin:
            if process_admin[chat_id] == add_lesson_date:
                try:
                    day, month, year = map(int, text.split("."))
                    book = openpyxl.load_workbook(filename)
                    sheet = book[main.sheetname]
                    for row in sheet.iter_rows():
                        for title in row:
                            if sheet[find_next_column(title.column_letter) + "1"].value is None:
                                sheet[find_next_column(title.column_letter) + "1"].value = datetime.date(year, month, day)
                                book.save(filename)
                                send_message(chat_id, Params(added_success))
                                process_admin.pop(chat_id)
                                break
                        break
                except Exception as e:
                    send_message(chat_id, Params(incorrect_date_format))
            elif len(process_admin[chat_id]) == 2 and process_admin[chat_id][1] == choose_date:
                day, month, year = text.split("-")
                if mark_user(0, surname=process_admin[chat_id][0], current_data=datetime.date(int(year), int(month), int(day))):
                    send_message(chat_id,
                                 Params(text=student_marked, reply_markup=create_reply_markup(admin_tools)))
                else:
                    send_message(chat_id, Params(text=student_unmarked,
                                                 reply_markup=create_reply_markup(admin_tools)))
                process_admin.pop(chat_id)
        else:
            send_message(chat_id, Params(text=else_admin_command))
    elif message.is_document_message() and message.get_file_type() == "xlsx":
        try:
            main_book = openpyxl.load_workbook(filename)
            new_table = get_new_book(message)
            main_sheet = main_book.create_sheet(main.sheetname)
            for i in range(len(new_table)):
                for j in new_table[i][:3]:
                    main_sheet[j.coordinate].value = j.value
            main_sheet.insert_cols(0)
            main_sheet["A1"].value = "chat id"
            main_sheet["D1"].value = "Достижения"
            put_dates(main_sheet)
            main_book.save(filename)
            send_message(chat_id, Params(text=added_success))
        except Exception as e:
            send_message(chat_id, Params(text=f'Ошибка {e}'))


class NewPerson:
    def __init__(self):
        self.surname = None
        self.course = None
        self.advantage = None

    def get_surname(self):
        return self.surname

    def get_course(self):
        return self.course

    def get_advantage(self):
        return self.advantage

    def set_surname(self, surname):
        self.surname = surname

    def set_course(self, course):
        self.course = course

    def set_advantage(self, advantage):
        self.advantage = advantage
