import datetime

import pymorphy2
import requests
import time
from math import cos, sin, asin, sqrt, radians
import time
import String
import admin
from message import Message
import _
import openpyxl
import json
import schedule
from admin import *
from params import Params
from String import *
import threading

"""При перезапуске сервера всегда будет лист 2022 - 2023 (2)"""
TOKEN = _.token
sheetname = "2022 - 2023 (2)"
chat_id_column = "A"
course_column = "B"
surname_column = "C"
advantage_column = "D"
URL = 'https://api.telegram.org/bot'
radius = 0.3
lesson = pymorphy2.MorphAnalyzer().parse("занятие")[0]
center = 59.965128, 30.398474

process = {}


def get_updates(offset=0):
    result = requests.get(f'{URL + TOKEN}/getUpdates?offset={offset}').json()
    if "result" in result: return result['result']


def send_message(chat_id, params):
    requests.get(f'{URL + TOKEN}/sendMessage?chat_id={chat_id}', params=params.get_params())


def get_person_remoteness(location):
    if location is not None:
        center_point = [{latitude: center[0], longitude: center[1]}]
        lat1 = center_point[0][latitude]
        lon1 = center_point[0][longitude]
        lat2 = location[latitude]
        lon2 = location[longitude]
        lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
        dlon = lon2 - lon1
        dlat = lat2 - lat1
        a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
        c = 2 * asin(sqrt(a))
        r = 6371
        return c * r


# def mark_user(chat_id, surname=None, current_data=None):
#     today_data = datetime.date.today()
#     book = openpyxl.load_workbook(filename)
#     sheet = book[sheetname]
#     res_row = 0
#     for row in sheet.iter_rows():
#         if row[0].value == chat_id or row[2].value == surname:
#             res_row = str(row[0].row)
#             break
#     for row in sheet.iter_rows():
#         for title in row:
#             if isinstance(title.value, datetime.datetime):
#                 data = datetime.date(title.value.year, title.value.month, title.value.day)
#                 if (data == today_data and current_data is None) or data == current_data and res_row != "0":
#                     sheet[title.column_letter + res_row].value = "+"
#                     book.save(filename)
#                     return True
#         return False


def choose_surname(chat_id, text):
    book = openpyxl.load_workbook(filename)
    sheet = book[sheetname]
    selected_course = text.split()[0]
    persons = []
    for i in sheet[course_column][1:]:
        if str(i.value) == selected_course:
            persons.append(sheet[surname_column + str(i.row)].value)
    send_message(chat_id, Params(text=choose_surname_command,
                                 reply_markup=create_reply_markup(chunked_list(sorted(persons)))))


def chunked_list(list):
    result = []
    for i in range(0, len(list), 3):
        if i + 2 < len(list):
            result.append([list[i], list[i + 1], list[i + 2]])
    if len(list) % 3 == 1:
        result.append([list[-1]])
    if len(list) % 3 == 2:
        result.append([list[-2], list[-1]])
    return result


def write_logs(message):
    date = datetime.datetime.utcfromtimestamp(message.get_date()) + datetime.timedelta(hours=3)
    file = open(archive, "a+")
    try:
        file.write(f"chatId - {message.get_chat_id()}; username - {message.get_username()}; "
                   f"date {date}; text - {message.get_text()}; location - "
                   f"{message.is_live_location()}; radius - {get_person_remoteness(message.get_location())}" + "\n")
    finally:
        file.close()


def get_attendance(chat_id):
    book = openpyxl.open(filename)[sheetname]
    all_chat_ids = [i.value for i in book[chat_id_column][1:]]
    chat_id_index = all_chat_ids.index(chat_id) + 2
    attendances_date = []
    for i in book.iter_rows():
        for title in i:
            if isinstance(title.value, datetime.datetime):
                if book[title.column_letter + str(chat_id_index)].value == "+":
                    attendances_date.append(title.value)
        break
    return attendances_date


def mark_user(chat_id, surname=None, current_data=None):
    today_data = datetime.date.today()
    book = openpyxl.load_workbook(filename)
    sheet = book[sheetname]
    res_row = 0
    for row in sheet.iter_rows():
        if row[0].value == chat_id or row[2].value == surname:
            res_row = str(row[0].row)
            break
    for row in sheet.iter_rows():
        for title in row:
            if isinstance(title.value, datetime.datetime):
                data = datetime.date(title.value.year, title.value.month, title.value.day)
                if (data == today_data and current_data is None) or data == current_data and res_row != "0":
                    sheet[title.column_letter + res_row].value = "+"
                    book.save(filename)
                    return True
        return False


def find_next_column(n):
    if isinstance(n, str):
        if n >= "Z":
            return "A" + chr(ord(n) - 25)
        else:
            return chr(ord(n) + 1)
    elif isinstance(n, int):
        if n >= 26:
            return "A" + chr(ord("A") + n % 26)
        else:
            return chr(ord("A") + n)


def create_reply_markup(list):
    return json.dumps({'keyboard': list,
                       'resize_keyboard': True,
                       'one_time_keyboard': True,
                       'selective': True})


def is_user_authorization(chat_id):
    book = openpyxl.load_workbook(filename)
    return chat_id in [i.value for i in book[sheetname][chat_id_column]][1:]


def get_surname_by_chat_id(chat_id):
    book = openpyxl.load_workbook(filename)
    sheet = book[sheetname]
    chat_ids = [i.value for i in book[sheetname][chat_id_column]][1:]
    return sheet[surname_column + str(2 + chat_ids.index(chat_id))].value


def get_chat_id_field_by_surname(surname):
    book = openpyxl.load_workbook(filename)
    sheet = book[sheetname]
    all_surname = [i.value for i in sheet[surname_column][1:]]
    surname_index = all_surname.index(surname)
    return sheet[chat_id_column + str(2 + surname_index)]


def is_surname_exist(surname):
    book = openpyxl.load_workbook(filename)
    sheet = book[sheetname]
    return surname in [i.value for i in sheet[surname_column][1:]]


def get_attendance_for_user(chat_id):
    attendance = get_attendance(chat_id)
    if len(attendance) == 0:
        send_message(chat_id, Params("Вас не было ни на одном занятии"))
        send_media(chat_id, sad_pic, "photo")
    else:
        attendance = list(
            map(lambda x: f"   - {str(x.day)}.{str(x.month)}.{str(x.year)}",
                attendance))
        send_message(chat_id, Params("\n".join([f"Вы посетили {len(attendance)} "
                                                f"{lesson.make_agree_with_number(len(attendance)).word}:"] + attendance)))


def start(message):
    chat_id = message.get_chat_id()
    if message.is_text_message():
        if start_command == message.get_text():
            send_message(chat_id, Params(text=choose_course,
                                         reply_markup=create_reply_markup([[first_course, second_course]])))
        elif course in message.get_text():
            choose_surname(chat_id, message.get_text())
        else:
            book = openpyxl.load_workbook(filename)
            sheet = book[sheetname]
            if is_surname_exist(message.get_text()):
                chat_id_field = get_chat_id_field_by_surname(message.get_text())
                if not str(chat_id_field.value).isdigit():
                    sheet[chat_id_field.coordinate].value = chat_id
                    book.save(filename)
                    reply_markup = json.dumps({remove_keyboard: True})
                    send_message(chat_id, Params(text=authorization_success, reply_markup=reply_markup))
                    send_message(chat_id, Params(follow_instructions))
                    send_media(chat_id, requirements_of_location_message, "photo")
                else:
                    send_message(chat_id, Params(text=surname_used_by_other))
            else:
                send_message(chat_id, Params(text=surname_enter_incorrect))


# def check_user(chat_id):
#     if chat_id in process:
#         file = open("live_location.txt", "a+")
#         try:
#             file.write(f"{chat_id}: {str(process[chat_id])} \n")
#             schedule.clear(chat_id)
#         finally:
#             file.close()


def run():
    update_id = get_updates()[-1]["update_id"]
    while True:
        schedule.run_pending()
        try:
            messages = get_updates(update_id)
            time.sleep(1)
            for message_js in messages:
                if update_id < message_js["update_id"]:
                    update_id = message_js["update_id"]
                    message = Message(message_js)

                    if message.get_message_type() == "message":
                        chat_id = message.get_chat_id()
                        write_logs(message)
                        if message.get_username() in admins:
                            admin_method(message)
                        elif not is_user_authorization(chat_id):
                            start(message)
                        elif message.is_location_message():
                            if message.is_live_location():
                                if get_person_remoteness(message.get_location()) <= radius:
                                    # process[chat_id] = []
                                    if mark_user(chat_id):
                                        # schedule.every(62).minutes.do(check_user, chat_id=chat_id).tag(chat_id)
                                        send_message(chat_id, Params(text=correct_input))
                                    else:
                                        reply_markup = json.dumps({remove_keyboard: True})
                                        send_message(chat_id, Params(text=no_lesson_today,
                                                                     reply_markup=reply_markup))
                                        send_media(chat_id, sad_pic, "photo")
                                else:
                                    send_message(chat_id, Params(
                                        text=f"Для отметки о посещении необходимо быть в радиусе "
                                             f"{int(radius * 1000)} метров"))
                            else:
                                send_message(chat_id, Params(text=notification_of_location))
                                send_media(chat_id, requirements_of_location_message, "photo")
                        else:
                            if message.get_text() == "file" and chat_id == 5084780807:
                                send_media(chat_id, archive, "document")
                            elif message.get_text() == "live" and chat_id == 5084780807:
                                send_media(chat_id, "logs/live_location.txt", "document")

                            elif message.get_text() == str_get_attendance_user:
                                get_attendance_for_user(chat_id)
                            else:
                                send_message(chat_id, Params(
                                    text="Вы вошли под фамилией " + get_surname_by_chat_id(chat_id)))
                        # elif message.get_message_type() == "edited_message":
                        #     chat_id = message.get_chat_id()
                        #     if chat_id in process and message.is_live_location():
                        #         date = datetime.datetime.utcfromtimestamp(message.get_edit_date()) + datetime.timedelta(
                        #             hours=3)
                        #         process[chat_id].append([date, get_person_remoteness(message.get_location())])
        except Exception as e:
            print(e)
            time.sleep(1)


if __name__ == '__main__':
    main = threading.Thread(target=run)
    main.start()
